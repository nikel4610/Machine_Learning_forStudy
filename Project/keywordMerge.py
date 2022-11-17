import pandas as pd
import os
import openpyxl as xls
import glob

file_format = ".xlsx"
file_path = 'D:/vsc_project/machinelearning_study/Project/searchData/Data/Keyword_Count'
file_name_os = os.listdir(file_path)
# # print(file_name_os)

file_list = [f"{file_path}/{file}" for file in os.listdir(file_path) if file_format in file]
# print(file_list)

for i in range(len(file_name_os)):
    wb = xls.load_workbook(file_list[i])
    ws = wb.active
    ws.delete_cols(3, 2)
    wb.save(file_list[i])

for i in range(len(file_name_os)):
    wb = xls.load_workbook(file_list[i])
    ws = wb.active
    ws['C1'].value = ws['B2'].value
    wb.save(file_list[i])

for i in range(len(file_name_os)):
    wb = xls.load_workbook(file_list[i])
    ws = wb.active
    ws.delete_cols(2)
    wb.save(file_list[i])

for i in range(len(file_name_os)-1):
    wb = xls.load_workbook(file_list[i])
    ws = wb.active
    ws.delete_cols(1)
    wb.save(file_list[i])

df = pd.DataFrame()
for i in range(len(file_name_os)):
    df = pd.concat([df, pd.read_excel(file_list[i])], axis = 1)
df.to_excel(file_path + '/' + 'datalabAll' + '.xlsx')


